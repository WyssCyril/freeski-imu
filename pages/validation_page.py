"""
Validierung: Bland-Altman, Typical Error, Sensor-Vergleich (Cyril & Nils)
Datenquelle: Validierungsmessungen (Kraftmessplatte + IMU-Sensoren)
"""
import streamlit as st
import numpy as np
import pandas as pd
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from utils.statistics import bland_altman, plot_bland_altman, plot_correlation

VAL_DIR = ("/Users/cyrilwyss/AAMasterarbeit/Data/Validierungsmessungen/"
           "Datasheets Validierungsmessungen neu (Nils & Cyril)")
import os as _os
_VAL_DIR_EXISTS = _os.path.exists(VAL_DIR)

# Schwellenwerte
WARN_PCT  = 15.0
EXCL_PCT  = 25.0

# Sensor-Definition: Anzeigename → (IMU-Spalte, Referenz-Spalte)
SENSORS = {
    "Bauch 1 (b1)":        ("peak_res_g_b1",  "peak_res_g_KMP"),
    "Bauch 2 (b2)":        ("peak_res_g_b2",  "peak_res_g_KMP"),
    "Fuss links 1 (li1)":  ("peak_res_g_li1", "peak_res_g_KMP"),
    "Fuss links 2 (li2)":  ("peak_res_g_li2", "peak_res_g_KMP"),
    "Fuss rechts 1 (re1)": ("peak_res_g_re1", "peak_res_g_KMP"),
    "Fuss rechts 2 (re2)": ("peak_res_g_re2", "peak_res_g_KMP"),
}

# Für b1/b2 sind auch weitere Parameter verfügbar
EXTRA_PARAMS = {
    "Peak-g":       ("peak_res_g",     "peak_res_g_KMP"),
    "RFD (g/s)":    ("RFD_to_peak",    None),
    "Impuls (g·s)": ("Impulse_to_peak_net", None),
    "Time to Peak": ("time_to_peak",   None),
}


def _nils_to_cyril_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Normalisiert Nils-Spaltennamen auf das Cyril-Format (_b1/_b2/etc.)"""
    rename = {
        # Bauch 1
        "takeoff_idx":          "takeoff_idx_b1",
        "landing_idx":          "landing_idx_b1",
        "flight_time_s":        "flight_time_s_b1",
        "time_to_peak":         "time_to_peak_b1",
        "peak_vert_idx":        "peak_vert_idx_b1",
        "peak_vert_g":          "peak_vert_g_b1",
        "peak_res_idx":         "peak_res_idx_b1",
        "peak_res_timestamp":   "peak_res_timestamp_b1",
        "peak_res_g":           "peak_res_g_b1",
        "RFD_to_peak":          "RFD_to_peak_b1",
        "Impulse_to_peak":      "Impulse_to_peak_b1",
        "Impulse_to_peak_net":  "Impulse_to_peak_net_b1",
        "landing_posix_time":   "landing_posix_time_b1",
        "landing_local_time":   "landing_local_time_b1",
        # Bauch 2
        "takeoff_idx.1":          "takeoff_idx_b2",
        "landing_idx.1":          "landing_idx_b2",
        "flight_time_s.1":        "flight_time_s_b2",
        "time_to_peak.1":         "time_to_peak_b2",
        "peak_vert_idx.1":        "peak_vert_idx_b2",
        "peak_vert_g.1":          "peak_vert_g_b2",
        "peak_res_idx.1":         "peak_res_idx_b2",
        "peak_res_timestamp.1":   "peak_res_timestamp_b2",
        "peak_res_g.1":           "peak_res_g_b2",
        "RFD_to_peak.1":          "RFD_to_peak_b2",
        "Impulse_to_peak.1":      "Impulse_to_peak_b2",
        "Impulse_to_peak_net.1":  "Impulse_to_peak_net_b2",
        "landing_posix_time.1":   "landing_posix_time_b2",
        "landing_local_time.1":   "landing_local_time_b2",
        # Fuss-Sensoren
        "sensor_place.1":       "sensor_place_li1",
        "peak_res_g.2":         "peak_res_g_li1",
        "landing_local_time.2": "landing_local_time_li1",
        "sensor_place.2":       "sensor_place_re1",
        "peak_res_g.3":         "peak_res_g_re1",
        "landing_local_time.3": "landing_local_time_re1",
        "sensor_place.3":       "sensor_place_re2",
        "peak_res_g.4":         "peak_res_g_re2",
        "landing_local_time.4": "landing_local_time_re2",
    }
    return df.rename(columns=rename)


@st.cache_data
def load_validation_data() -> pd.DataFrame:
    """Lädt und kombiniert Cyril + Nils Validierungsdaten."""
    path_c = f"{VAL_DIR}/Validierung Cyril.xlsx"
    path_n = f"{VAL_DIR}/Validierung Nils.xlsx"

    df_c = pd.read_excel(path_c, sheet_name="Validierung")
    df_n = pd.read_excel(path_n, sheet_name="Validierung")
    df_n = _nils_to_cyril_cols(df_n)

    # peak_res_g_KMP für Nils berechnen (fehlt im Validierungssheet)
    if "peak_res_g_KMP" not in df_n.columns or df_n["peak_res_g_KMP"].isna().all():
        df_n["peak_res_g_KMP"] = df_n["peak_landing_F"] / (df_n["body_mass"] * 9.81)

    df = pd.concat([df_c, df_n], ignore_index=True)
    df = df[df["exercise"].notna()]
    return df


def _cv_label(cv: float) -> str:
    if cv < 10:  return "gut (< 10 %)"
    if cv < 20:  return "akzeptabel (10–20 %)"
    return "ungenügend (> 20 %)"


def _render_report(ba: dict, ba_per_ath: dict, overview_rows: list,
                   sensor: str, athlete: str, exercise: str,
                   x_all: np.ndarray, y_all: np.ndarray,
                   excl_indices: list, manual_reasons: dict,
                   df_meta: pd.DataFrame,
                   warn_pct: float, excl_pct: float):
    """
    Vollständiger Validierungsbericht:
    - Ausgeschlossene Datenpunkte mit Begründung
    - Korrekturempfehlung + Anwendung
    - Vorher / Nachher Bland-Altman
    """
    from utils.statistics import bland_altman, plot_bland_altman, plot_correlation
    from scipy import stats as sp_stats

    # ── Arbeitsdaten (ohne manuell ausgeschlossene) ────────────────────────
    mask_excl = np.zeros(len(x_all), dtype=bool)
    mask_excl[excl_indices] = True
    x_in = x_all[~mask_excl]
    y_in = y_all[~mask_excl]
    ba = bland_altman(x_in, y_in)

    grand_mean = float(np.mean(ba["mean"])) if len(ba["mean"]) else 1.0
    bias       = ba["bias"]
    cv         = ba["cv_pct"]
    te         = ba["typical_error"]
    loa_upper  = ba["loa_upper"]
    loa_lower  = ba["loa_lower"]
    r          = ba["r"]
    trend_sig  = ba["p_trend"] < 0.05
    slope      = ba["slope"]
    intercept  = ba["intercept"]
    n_in       = ba["n"]

    pct_devs_in = np.abs(ba["diff"] - bias) / grand_mean * 100 if grand_mean else np.zeros(n_in)
    n_warn      = int((pct_devs_in >= warn_pct).sum())
    n_excl_auto = int((pct_devs_in >= excl_pct).sum())

    ath_str = athlete if athlete != "Alle" else "Cyril & Nils"
    r_label = ("sehr stark" if abs(r) >= 0.9 else "stark" if abs(r) >= 0.7
               else "mässig" if abs(r) >= 0.5 else "schwach")
    bias_pct = abs(bias / grand_mean * 100) if grand_mean else 0

    # ══════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.subheader("📋 Validierungsbericht")

    # ── 1. Kontext ─────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown(f"**Sensor:** {sensor} &nbsp;|&nbsp; **Athlet:** {ath_str} &nbsp;|&nbsp; "
                    f"**Übung:** {exercise} &nbsp;|&nbsp; "
                    f"**n = {n_in}** Messwerte ({len(excl_indices)} manuell ausgeschlossen, {len(x_all)} total)")

    # ── 2. Ausgeschlossene Datenpunkte ─────────────────────────────────────
    st.markdown("#### Ausgeschlossene Datenpunkte")

    all_excluded_rows = []

    # Manuell ausgeschlossene
    for idx in excl_indices:
        if idx >= len(x_all):
            continue
        trial = df_meta["trial"].iloc[idx] if idx < len(df_meta) else "—"
        ath   = df_meta["athlete_id"].iloc[idx] if idx < len(df_meta) else "—"
        xi, yi = float(x_all[idx]), float(y_all[idx])
        dev_pct = abs(yi - xi) / abs(xi) * 100 if xi != 0 else 0
        reason_txt = manual_reasons.get(str(idx), "Manuell ausgeschlossen")
        all_excluded_rows.append({
            "Index": idx, "Trial": trial, "Athlet": ath,
            "KMP (g)": round(xi, 3), "IMU (g)": round(yi, 3),
            "Abweichung": f"{dev_pct:.1f} %",
            "Typ": "✋ Manuell",
            "Begründung": reason_txt,
        })

    # Automatisch: über Ausschlussschwelle in verbleibenden Daten
    pct_devs_all = np.abs((y_all - x_all) - bias) / grand_mean * 100 if grand_mean else np.zeros(len(x_all))
    for i, (xi, yi, pct) in enumerate(zip(x_all, y_all, pct_devs_all)):
        if i in excl_indices:
            continue
        if pct >= excl_pct:
            trial = df_meta["trial"].iloc[i] if i < len(df_meta) else "—"
            ath   = df_meta["athlete_id"].iloc[i] if i < len(df_meta) else "—"
            dev_pct = abs(yi - xi) / abs(xi) * 100 if xi != 0 else 0
            all_excluded_rows.append({
                "Index": i, "Trial": trial, "Athlet": ath,
                "KMP (g)": round(float(xi), 3), "IMU (g)": round(float(yi), 3),
                "Abweichung": f"{dev_pct:.1f} %",
                "Typ": f"⚠️ Auto (>{excl_pct:.0f} %)",
                "Begründung": f"Abweichung {dev_pct:.1f} % überschreitet Ausschlussschwelle {excl_pct:.0f} %",
            })

    if all_excluded_rows:
        st.dataframe(pd.DataFrame(all_excluded_rows), use_container_width=True, hide_index=True)
        st.caption(f"Schwellenwerte: Warnung ≥ {warn_pct:.0f} %, Ausschluss ≥ {excl_pct:.0f} %")
    else:
        st.success(f"Alle {n_in} Messwerte liegen innerhalb der Schwellenwerte — kein Ausschluss.")

    # ── 3. Befunde ─────────────────────────────────────────────────────────
    st.markdown("#### Befunde")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f"""
**Systematische Abweichung (Bias)**
Der IMU-Sensor {'überschätzt' if bias > 0 else 'unterschätzt'} die Kraftmessplatte
um **{bias:+.3f} g ({bias_pct:.1f} %)**.
{'Diese Abweichung ist vernachlässigbar (< 5 %).' if bias_pct < 5 else
 'Eine systematische Korrektur ist empfehlenswert.' if bias_pct >= 15 else
 'Die Abweichung ist moderat und sollte bei der Interpretation berücksichtigt werden.'}

Übereinstimmungsgrenzen (LoA): **{loa_lower:.3f} g** bis **{loa_upper:.3f} g**
(95 % CI ±{ba['ci_loa']:.3f} g)
""")
    with col2:
        st.markdown(f"""
**Messvariabilität**
CV = **{cv:.1f} %** → {_cv_label(cv)}
Typical Error = {te:.4f} g

**Korrelation**
Pearson r = **{r:.3f}** ({r_label})
{'⚠️ Proportionaler Bias (p = ' + f"{ba['p_trend']:.4f}" + '): Abweichung nimmt mit höherer Last zu.' if trend_sig else
 f'Kein proportionaler Bias (p = {ba["p_trend"]:.3f}).'}
""")

    if len(ba_per_ath) == 2:
        st.markdown("**Cyril vs. Nils**")
        for ath, ba_a in ba_per_ath.items():
            st.markdown(f"- **{ath}**: Bias = {ba_a['bias']:+.3f} g, "
                        f"CV = {ba_a['cv_pct']:.1f} % ({_cv_label(ba_a['cv_pct'])}), "
                        f"r = {ba_a['r']:.3f}")

    # ── 4. Korrekturempfehlung & Anwendung ─────────────────────────────────
    st.markdown("#### Korrekturempfehlung")

    need_bias_corr  = bias_pct >= 5
    need_prop_corr  = trend_sig

    if not need_bias_corr and not need_prop_corr:
        st.success("Keine Korrektur notwendig — der Sensor kann ohne Anpassung verwendet werden.")
        y_corrected = y_in.copy()
        corr_desc = "Keine Korrektur"
    elif need_prop_corr:
        st.info(
            f"**Empfohlene Korrektur: Lineare Regression (proportionaler Bias)**\n\n"
            f"Da ein signifikanter proportionaler Bias vorliegt (slope = {slope:.4f}, p = {ba['p_trend']:.4f}), "
            f"wird eine lineare Korrektur empfohlen:\n\n"
            f"```\ny_korrigiert = y_IMU − (slope × mean + intercept_der_Differenz)\n"
            f"            = y_IMU − ({slope:.4f} × mean + {intercept:.4f})\n```\n\n"
            f"Dies korrigiert sowohl den konstanten Bias ({bias:+.3f} g) als auch die "
            f"lastabhängige Zunahme der Abweichung."
        )
        mean_in = (x_in + y_in) / 2
        y_corrected = y_in - (slope * mean_in + intercept)
        corr_desc = f"Lineare Korrektur: y − ({slope:.4f}×mean + {intercept:.4f})"
    else:
        st.info(
            f"**Empfohlene Korrektur: Bias-Subtraktion**\n\n"
            f"Der konstante Bias von {bias:+.3f} g ({bias_pct:.1f} %) wird von allen IMU-Messwerten subtrahiert:\n\n"
            f"```\ny_korrigiert = y_IMU − {bias:.4f}\n```"
        )
        y_corrected = y_in - bias
        corr_desc = f"Bias-Subtraktion: y − {bias:.4f} g"

    # ba_corr immer berechnen (für Excel-Export)
    ba_corr = bland_altman(x_in, y_corrected)
    fig_before = fig_after = fig_corr2 = None

    # ── 5. Vorher / Nachher Vergleich ──────────────────────────────────────
    if need_bias_corr or need_prop_corr:
        st.markdown("#### Vorher / Nachher Vergleich")

        # Kennzahlen-Tabelle
        comp_df = pd.DataFrame({
            "Kennzahl": ["Bias (g)", "CV (%)", "LoA untere Grenze (g)", "LoA obere Grenze (g)",
                         "Typical Error (g)", "Pearson r"],
            "Ohne Korrektur": [
                f"{ba['bias']:+.3f}", f"{ba['cv_pct']:.1f}",
                f"{ba['loa_lower']:.3f}", f"{ba['loa_upper']:.3f}",
                f"{ba['typical_error']:.4f}", f"{ba['r']:.3f}",
            ],
            "Mit Korrektur": [
                f"{ba_corr['bias']:+.3f}", f"{ba_corr['cv_pct']:.1f}",
                f"{ba_corr['loa_lower']:.3f}", f"{ba_corr['loa_upper']:.3f}",
                f"{ba_corr['typical_error']:.4f}", f"{ba_corr['r']:.3f}",
            ],
        })
        st.dataframe(comp_df, use_container_width=True, hide_index=True)

        fig_before = plot_bland_altman(ba,
            title=f"Bland-Altman: IMU {sensor} vs. Kraftmessplatte — Ohne Korrektur",
            warn_pct=warn_pct, excl_pct=excl_pct)
        fig_after = plot_bland_altman(ba_corr,
            title=f"Bland-Altman: IMU {sensor} vs. Kraftmessplatte — Mit Korrektur",
            warn_pct=warn_pct, excl_pct=excl_pct)
        fig_corr2 = plot_correlation(x_in, y_corrected,
            "Kraftmessplatte (g)", f"IMU {sensor} (korrigiert)",
            title=f"Korrelation: IMU {sensor} vs. Kraftmessplatte — Nach Korrektur")

        c_before, c_after = st.columns(2)
        with c_before:
            st.plotly_chart(fig_before, use_container_width=True)
        with c_after:
            st.plotly_chart(fig_after, use_container_width=True)
        st.plotly_chart(fig_corr2, use_container_width=True)

    # ── 6. Gesamtbeurteilung ───────────────────────────────────────────────
    st.markdown("#### Gesamtbeurteilung")
    if cv < 10 and bias_pct < 5 and not trend_sig:
        verdict_col = "success"
        verdict_txt = (
            f"✅ Der IMU-Sensor **{sensor}** misst sehr ähnlich wie die Kraftmessplatte. "
            f"Die Abweichung beträgt im Durchschnitt nur **{bias:+.3f} g ({bias_pct:.1f} %)** — "
            f"das ist vernachlässigbar. Der Sensor kann direkt im Feld eingesetzt werden, ohne Korrektur."
        )
    elif cv < 20 and bias_pct < 15:
        verdict_col = "warning"
        verdict_txt = (
            f"⚠️ Der IMU-Sensor **{sensor}** misst ähnlich wie die Kraftmessplatte, "
            f"weicht aber im Schnitt um **{bias:+.3f} g ({bias_pct:.1f} %)** ab. "
            f"Die Streuung der Messwerte beträgt **{cv:.1f} %** — das ist akzeptabel. "
            + (f"Eine Korrektur wird empfohlen, um die Genauigkeit zu verbessern. " if need_bias_corr or need_prop_corr else "")
            + f"Die Ergebnisse sollten mit Vorsicht interpretiert werden."
        )
    else:
        verdict_col = "error"
        verdict_txt = (
            f"❌ Der IMU-Sensor **{sensor}** weicht deutlich von der Kraftmessplatte ab "
            f"(Abweichung: **{bias:+.3f} g / {bias_pct:.1f} %**, Streuung: **{cv:.1f} %**). "
            f"Die Messwerte sind ohne Korrektur nicht direkt mit der Kraftmessplatte vergleichbar. "
            f"Eine Korrektur ist zwingend erforderlich."
        )

    getattr(st, verdict_col)(verdict_txt)

    # ── Export ─────────────────────────────────────────────────────────────
    excluded_txt = "\n".join(
        f"  {r['Trial']} ({r['Athlet']}): IMU={r['IMU (g)']} g, KMP={r['KMP (g)']} g, "
        f"Abweichung={r['Abweichung']}, {r['Typ']} — {r['Begründung']}"
        for r in all_excluded_rows
    ) or "  Keine"

    export_txt = f"""VALIDIERUNGSBERICHT
Sensor: {sensor} | Athlet: {ath_str} | Übung: {exercise}
n = {n_in} (total: {len(x_all)}, ausgeschlossen: {len(excl_indices)})

AUSGESCHLOSSENE DATENPUNKTE
{excluded_txt}

BEFUNDE
Bias: {bias:+.3f} g ({bias_pct:.1f} %)
LoA: {loa_lower:.3f} g bis {loa_upper:.3f} g
CV: {cv:.1f} % ({_cv_label(cv)})
Typical Error: {te:.4f} g
Pearson r: {r:.3f} ({r_label})
Proportionaler Bias: {'ja (p=' + f'{ba["p_trend"]:.4f}' + ')' if trend_sig else 'nein'}

KORREKTUR
{corr_desc}

GESAMTBEURTEILUNG
{verdict_txt.replace('**', '')}
"""
    # Excel-Export mit Grafiken
    import io as _io
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Validierungsbericht"

    # Hilfsfunktion: Zelle schreiben
    def cell(row, col, value, bold=False, bg=None):
        c = ws.cell(row=row, column=col, value=value)
        if bold:
            c.font = Font(bold=True)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        return c

    row = 1
    cell(row, 1, f"Validierungsbericht: IMU {sensor} vs. Kraftmessplatte", bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    cell(row, 1, f"Athlet: {ath_str}  |  Übung: {exercise}  |  n = {n_in}")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    # Kennzahlen
    cell(row, 1, "Kennzahl", bold=True, bg="DDDDDD")
    cell(row, 2, "Ohne Korrektur", bold=True, bg="DDDDDD")
    cell(row, 3, "Mit Korrektur", bold=True, bg="DDDDDD")
    row += 1
    kennzahlen = [
        ("Bias (g)",            f"{ba['bias']:+.3f}",            f"{ba_corr['bias']:+.3f}" if (need_bias_corr or need_prop_corr) else "—"),
        ("CV (%)",              f"{ba['cv_pct']:.1f}",           f"{ba_corr['cv_pct']:.1f}" if (need_bias_corr or need_prop_corr) else "—"),
        ("LoA obere Grenze (g)",f"{ba['loa_upper']:.3f}",        f"{ba_corr['loa_upper']:.3f}" if (need_bias_corr or need_prop_corr) else "—"),
        ("LoA untere Grenze (g)",f"{ba['loa_lower']:.3f}",       f"{ba_corr['loa_lower']:.3f}" if (need_bias_corr or need_prop_corr) else "—"),
        ("Typical Error (g)",   f"{ba['typical_error']:.4f}",    f"{ba_corr['typical_error']:.4f}" if (need_bias_corr or need_prop_corr) else "—"),
        ("Pearson r",           f"{ba['r']:.3f}",                f"{ba_corr['r']:.3f}" if (need_bias_corr or need_prop_corr) else "—"),
        ("Korrektur",           "Keine",                         corr_desc),
        ("Gesamtbeurteilung",   verdict_txt.replace("✅ ","").replace("⚠️ ","").replace("❌ ",""), ""),
    ]
    for k, v1, v2 in kennzahlen:
        cell(row, 1, k)
        cell(row, 2, v1)
        cell(row, 3, v2)
        row += 1
    row += 1

    # Rohdaten
    cell(row, 1, "Rohdaten", bold=True, bg="DDDDDD")
    cell(row, 2, "Kraftmessplatte (g)", bold=True, bg="DDDDDD")
    cell(row, 3, "IMU (g)", bold=True, bg="DDDDDD")
    cell(row, 4, "Differenz (g)", bold=True, bg="DDDDDD")
    cell(row, 5, "Mittelwert (g)", bold=True, bg="DDDDDD")
    row += 1
    for xi, yi in zip(x_in, y_in):
        ws.cell(row=row, column=2, value=round(float(xi), 4))
        ws.cell(row=row, column=3, value=round(float(yi), 4))
        ws.cell(row=row, column=4, value=round(float(yi - xi), 4))
        ws.cell(row=row, column=5, value=round(float((xi + yi) / 2), 4))
        row += 1

    # Spaltenbreiten
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 25

    # Grafiken als Bilder einfügen (neues Sheet)
    try:
        ws_plots = wb.create_sheet("Grafiken")
        img_row = 1
        plots_to_add = [
            (fig_before, "Bland-Altman ohne Korrektur") if (need_bias_corr or need_prop_corr) else None,
            (fig_after,  "Bland-Altman mit Korrektur")  if (need_bias_corr or need_prop_corr) else None,
            (fig_corr2,  "Korrelation nach Korrektur")  if (need_bias_corr or need_prop_corr) else None,
        ]
        for item in plots_to_add:
            if item is None:
                continue
            fig_obj, title = item
            img_bytes = fig_obj.to_image(format="png", width=700, height=450, scale=1.5)
            img_stream = _io.BytesIO(img_bytes)
            xl_img = XLImage(img_stream)
            xl_img.anchor = f"A{img_row}"
            ws_plots.add_image(xl_img)
            img_row += 25
    except Exception:
        pass  # kaleido nicht verfügbar — Grafiken werden weggelassen

    excel_buf = _io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)
    st.download_button(
        "📥 Bericht als Excel exportieren",
        data=excel_buf,
        file_name=f"Validierungsbericht_{sensor.replace(' ','_')}_{athlete}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _cv_light(cv: float) -> str:
    if cv < 10:  return "🟢"
    if cv < 20:  return "🟡"
    return "🔴"


# Sensor-Gruppen mit Positions-Emoji
SENSOR_GROUPS = {
    "🎽 Bauch":       ["Bauch 1 (b1)", "Bauch 2 (b2)"],
    "🦶 Fuss links":  ["Fuss links 1 (li1)", "Fuss links 2 (li2)"],
    "🦶 Fuss rechts": ["Fuss rechts 1 (re1)", "Fuss rechts 2 (re2)"],
}


def _render_sensor_panel(df: pd.DataFrame, df_all: pd.DataFrame,
                         sname: str, sel_athlete: str, sel_exercise: str):
    """Rendert Kennzahlen, Plots, Ausreisser-Verwaltung und Bericht für einen Sensor."""
    imu_col, ref_col = SENSORS[sname]
    meta_cols = [c for c in ["trial", "athlete_id"] if c in df.columns]
    df_valid = df[[ref_col, imu_col] + meta_cols].dropna(subset=[ref_col, imu_col]).reset_index(drop=True)

    if len(df_valid) < 4:
        st.info(f"Zu wenig Datenpunkte (n={len(df_valid)}) für **{sname}**.")
        return

    x = df_valid[ref_col].values.astype(float)
    y = df_valid[imu_col].values.astype(float)

    excl_key = f"excl_{sname}_{sel_athlete}_{sel_exercise}"
    if excl_key not in st.session_state:
        st.session_state[excl_key] = []

    mask_excl = np.zeros(len(x), dtype=bool)
    mask_excl[st.session_state[excl_key]] = True
    x_in, y_in = x[~mask_excl], y[~mask_excl]
    ba = bland_altman(x_in, y_in)

    # Kennzahlen
    st.caption(f"n = {ba['n']}  ({len(df_valid)} total, {mask_excl.sum()} ausgeschlossen)")
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Bias", f"{ba['bias']:.3f} g")
    m1.caption(f"95% CI ±{ba['ci_bias']:.3f}")
    m2.metric("+LoA", f"{ba['loa_upper']:.3f} g")
    m2.caption(f"±{ba['ci_loa']:.3f}")
    m3.metric("−LoA", f"{ba['loa_lower']:.3f} g")
    m3.caption(f"±{ba['ci_loa']:.3f}")
    m4.metric("CV%", f"{ba['cv_pct']:.1f}% {_cv_light(ba['cv_pct'])}")
    m4.caption(f"TE = {ba['typical_error']:.4f}")
    m5.metric("Pearson r", f"{ba['r']:.3f}")

    if ba["p_trend"] < 0.05:
        st.warning(f"Proportionaler Bias (p = {ba['p_trend']:.4f})")

    # Plots
    p_ba, p_corr = st.tabs(["Bland-Altman", "Korrelation"])
    with p_ba:
        st.plotly_chart(
            plot_bland_altman(ba,
                title=f"Bland-Altman: IMU {sname} vs. Kraftmessplatte",
                warn_pct=WARN_PCT, excl_pct=EXCL_PCT),
            use_container_width=True)
    with p_corr:
        st.plotly_chart(
            plot_correlation(x_in, y_in,
                             x_label="Kraftmessplatte (g)", y_label=f"IMU {sname} (g)",
                             title=f"Korrelation: IMU {sname} vs. Kraftmessplatte"),
            use_container_width=True)

    # Ausreisser
    grand_mean = float(np.mean(ba["mean"])) if len(ba["mean"]) else 1.0
    ba_full = bland_altman(x, y)
    pct_devs_all = np.abs(ba_full["diff"] - ba_full["bias"]) / grand_mean * 100 if grand_mean else np.zeros(len(x))

    outlier_rows = []
    for i, (xi, yi, pct) in enumerate(zip(x, y, pct_devs_all)):
        if pct >= WARN_PCT or i in st.session_state[excl_key]:
            status = "⛔ Ausgeschlossen" if i in st.session_state[excl_key] else \
                     ("⛔ Ausschluss" if pct >= EXCL_PCT else "🟡 Warnung")
            outlier_rows.append({
                "Index": i,
                "Trial": df_valid["trial"].iloc[i] if "trial" in df_valid.columns else "—",
                "Athlet": df_valid["athlete_id"].iloc[i] if "athlete_id" in df_valid.columns else "—",
                "Referenz (g)": round(xi, 3),
                "IMU (g)": round(yi, 3),
                "% Abweichung": round(pct, 1),
                "Status": status,
            })

    if outlier_rows:
        with st.expander(f"Ausreisser ({len(outlier_rows)})", expanded=False):
            st.dataframe(pd.DataFrame(outlier_rows), use_container_width=True, hide_index=True)
            excl_new = st.multiselect("Manuell ausschliessen (Index):",
                                      options=list(range(len(x))),
                                      default=st.session_state[excl_key],
                                      key=f"ms_{excl_key}")
            reason = st.text_input("Begründung:", key=f"r_{excl_key}")
            if st.button("Speichern", key=f"btn_{excl_key}"):
                st.session_state[excl_key] = excl_new
                if reason:
                    st.session_state[f"{excl_key}_reason"] = reason
                st.rerun()

    # Cyril vs. Nils
    athletes_avail = df_all["athlete_id"].dropna().unique().tolist()
    ba_per_ath = {}
    if "Cyril" in athletes_avail and "Nils" in athletes_avail:
        with st.expander("Cyril vs. Nils", expanded=False):
            col_c, col_n = st.columns(2)
            for ath, col in [("Cyril", col_c), ("Nils", col_n)]:
                df_a = df_all[df_all["athlete_id"] == ath]
                if sel_exercise != "Alle":
                    df_a = df_a[df_a["exercise"] == sel_exercise]
                sub = df_a[[ref_col, imu_col]].dropna()
                if len(sub) < 4:
                    col.info(f"{ath}: n={len(sub)} (zu wenig)")
                    continue
                ba_a = bland_altman(sub[ref_col].values.astype(float),
                                    sub[imu_col].values.astype(float))
                ba_per_ath[ath] = ba_a
                col.plotly_chart(
                    plot_bland_altman(ba_a, title=f"{ath} | {sname}",
                                      warn_pct=WARN_PCT, excl_pct=EXCL_PCT),
                    use_container_width=True)
                col.caption(f"n={ba_a['n']} | Bias={ba_a['bias']:.3f} g | "
                            f"CV={ba_a['cv_pct']:.1f}% {_cv_light(ba_a['cv_pct'])}")

    # Bericht
    manual_reasons = {}
    reason_txt = st.session_state.get(f"{excl_key}_reason", "")
    if reason_txt:
        for idx in st.session_state[excl_key]:
            manual_reasons[str(idx)] = reason_txt

    _render_report(
        ba, ba_per_ath, [],
        sname, sel_athlete, sel_exercise,
        x_all=x, y_all=y,
        excl_indices=st.session_state[excl_key],
        manual_reasons=manual_reasons,
        df_meta=df_valid,
        warn_pct=WARN_PCT, excl_pct=EXCL_PCT,
    )


def show():
    st.header("Validierung")
    st.caption("Kraftmessplatte (Referenz) vs. IMU-Sensoren — Drop Jump Landings, Magglingen 2026")

    try:
        df_all = load_validation_data()
    except FileNotFoundError as e:
        st.error(f"Validierungsdatei nicht gefunden: {e}")
        return
    except Exception as e:
        st.error(f"Fehler beim Laden: {e}")
        return

    # ── Globale Filter ────────────────────────────────────────────────────
    c1, c2 = st.columns(2)
    athletes  = ["Alle"] + sorted(df_all["athlete_id"].dropna().unique().tolist())
    exercises = ["Alle"] + sorted(df_all["exercise"].dropna().unique().tolist())
    sel_athlete  = c1.selectbox("Athlet", athletes)
    sel_exercise = c2.selectbox("Übung",  exercises)

    df = df_all.copy()
    if sel_athlete  != "Alle": df = df[df["athlete_id"] == sel_athlete]
    if sel_exercise != "Alle": df = df[df["exercise"]   == sel_exercise]

    # ── Körperposition-Tabs ───────────────────────────────────────────────
    group_tabs = st.tabs(list(SENSOR_GROUPS.keys()) + ["📊 Überblick"])

    for tab, (group_name, sensor_names) in zip(group_tabs[:-1], SENSOR_GROUPS.items()):
        with tab:
            # Sensor-Vergleich: welcher korreliert besser?
            if len(sensor_names) == 2:
                stats = {}
                for sname in sensor_names:
                    imu_col, ref_col = SENSORS[sname]
                    sub = df[[ref_col, imu_col]].dropna()
                    if len(sub) >= 4:
                        ba_s = bland_altman(sub[ref_col].values.astype(float),
                                            sub[imu_col].values.astype(float))
                        stats[sname] = ba_s

                if len(stats) == 2:
                    s1, s2 = sensor_names
                    r1, r2 = stats[s1]["r"], stats[s2]["r"]
                    cv1, cv2 = stats[s1]["cv_pct"], stats[s2]["cv_pct"]
                    better = s1 if abs(r1) >= abs(r2) else s2
                    worse  = s2 if better == s1 else s1
                    r_b = stats[better]["r"]
                    r_w = stats[worse]["r"]
                    st.info(
                        f"**Sensor-Vergleich:** **{better}** korreliert besser mit der Kraftmessplatte "
                        f"(r = {r_b:.3f}) als **{worse}** (r = {r_w:.3f}). "
                        f"{'Der Unterschied ist gering.' if abs(r_b - r_w) < 0.05 else 'Der Unterschied ist deutlich.'}"
                    )

                for sname in sensor_names:
                    st.subheader(sname)
                    _render_sensor_panel(df, df_all, sname, sel_athlete, sel_exercise)
                    st.divider()
            else:
                for sname in sensor_names:
                    st.subheader(sname)
                    _render_sensor_panel(df, df_all, sname, sel_athlete, sel_exercise)

    # ── Überblick-Tab ─────────────────────────────────────────────────────
    with group_tabs[-1]:
        st.subheader("Alle Sensoren im Überblick")
        rows = []
        for sname, (icol, rcol) in SENSORS.items():
            sub = df[[rcol, icol]].dropna()
            if len(sub) < 4:
                continue
            ba_s = bland_altman(sub[rcol].values.astype(float), sub[icol].values.astype(float))
            # Gruppe bestimmen
            gruppe = next((g for g, names in SENSOR_GROUPS.items() if sname in names), "—")
            rows.append({
                "Gruppe": gruppe,
                "Sensor": sname,
                "n": ba_s["n"],
                "Bias (g)": f"{ba_s['bias']:.3f}",
                "+LoA": f"{ba_s['loa_upper']:.3f}",
                "−LoA": f"{ba_s['loa_lower']:.3f}",
                "CV%": f"{ba_s['cv_pct']:.1f}% {_cv_light(ba_s['cv_pct'])}",
                "TE": f"{ba_s['typical_error']:.4f}",
                "r": f"{ba_s['r']:.3f}",
                "Trend": "⚠️" if ba_s["p_trend"] < 0.05 else "✓",
            })
        if rows:
            st.dataframe(
                pd.DataFrame(rows).sort_values("Gruppe"),
                use_container_width=True, hide_index=True,
            )
